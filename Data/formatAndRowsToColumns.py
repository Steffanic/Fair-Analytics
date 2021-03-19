'''
Author: Patrick Steffanic 
Year: 2021
Purpose: The strategy here is to open each year's STOP report and format each class into a new sheet with Features in the columns and samples(fairs) as rows. 
'''


from openpyxl import load_workbook

DEBUGLVL = 0


def yield_excel_cols(cond_col):
    '''
    This function takes a column label and returns the next column.
    It assumes that you need 2 characters at most, e.g. "AF", "BB". This covers 26 + 26^2 = 702 columns.
    We have o.t.o.o. 100 rows to transpose into columns so this should be sufficient.
    To scale this function would require some thought about alphabet arithemtic.
    '''
    new_col = ""
    if(len(cond_col) == 1):
        if(str(chr(ord(cond_col[0])+1)) == "["):
            new_col = "AA"
        else:
            new_col = str(chr(ord(cond_col[0])+1))
    else:
        if(str(chr(ord(cond_col[-1])+1)) == "["):
            new_col = str(chr(ord(cond_col[0])+1)) + "A"
        else:
            new_col = cond_col[0] + str(chr(ord(cond_col[-1])+1))

    return new_col


def format_fair_to_new_sheet(col_label, new_row, Cur_Class, Cur_Class_FSheet):
    cond_col = "A"  # we always start at column A
    for row in range(1, 200):  # This iterates through the rows of the given fair
        if not all(v is None for v in [Cur_Class[f'{chr(ord("C")+x)}{row}'].value for x in range(8)]):  # This checks if the row is empty
            # or if it is the secondary name row
            if (row > 10 and (0 if type(Cur_Class[f'{col_label}{row}'].value) != str else (Cur_Class[f'{col_label}{row}'].value[:10] == Cur_Class[f'{col_label}{2}'].value[:10]))):
                continue
            # Assign the appropriate row value to the appropriate column within the new sheet
            Cur_Class_FSheet[f"{cond_col}{new_row}"] = Cur_Class[f'{col_label}{row}'].value
            cond_col = yield_excel_cols(cond_col)  # Advance the column cursor


def format_feature_labels(Cur_Class, Cur_Class_FSheet):
    cond_col = "A"
    # We iterate through all the rows to extract the appropriate feature names
    for row in range(1, 200):
        if (not all(v is None for v in [Cur_Class[f'{chr(ord("C")+x)}{row}'].value for x in range(8)])):
            # or if it is the secondary name row
            if (row > 10 and (Cur_Class[f'C{row}'].value == Cur_Class[f'C{2}'].value)):
                continue
            Cur_Class_FSheet[f"{cond_col}1"] = Cur_Class[f"A{row}"].value if Cur_Class[
                f"A{row}"].value != None else Cur_Class[f"B{row}"].value
            cond_col = yield_excel_cols(cond_col)
    Cur_Class_FSheet["A1"] = "Fair Name"


def format_workbook(year):

    print(f"Processing year {year}")

    # Load up the year workbook and use the values calculated from the last time the workbook was opened(Rather than maintaining the formulas within each cell.)
    STOP = load_workbook(filename=f"{year} STOP Summary.xlsx", data_only=True)

    classes = list(map(str, range(1, 8)))
    if int(year) >= 2009:
        classes.append("3+")
        classes.append("4+")

    for class_num in classes:

        print(f"Formatting class {class_num}")

        Cur_Class = STOP[f"Class {class_num}"]

        Cur_Class_FSheet = STOP.create_sheet(f"Class {class_num} Formatted")

        format_feature_labels(Cur_Class, Cur_Class_FSheet)

        col = "C"
        for row in range(2, 30):
            if Cur_Class[f"{col}2"].value == None:
                if DEBUGLVL == 1:
                    print(f"Finished on column {col}")
                break
            format_fair_to_new_sheet(col, row, Cur_Class, Cur_Class_FSheet)
            col = yield_excel_cols(col)

    # Better save if we want to see our changes
    STOP.save(f'{year} STOP Summary.xlsx')

for year in range(2001, 2019):
    if year == 2011:
        continue
    format_workbook(year)

