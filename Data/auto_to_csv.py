'''
Author: Patrick Steffanic 
Year: 2021
Purpose: The strategy here is to open each year's STOP report and format each class into a new sheet with Features in the columns and samples(fairs) as rows. 
'''


from openpyxl import load_workbook

STOP_2001 = load_workbook(filename="2001 STOP Summary.xlsx", data_only=True) # Load up the 2001 workbook and use the values calculated from the last time the workbook was opened(Rather than maintaining the formulas within each cell.)


Class_1 = STOP_2001["Class 1"] # Grab the first class of fairs

for row in range(1,50): # Cycle through the first fair to ensure that there are no Nones ^^This is a great place for input validation post-mvp^^
    if (Class_1[f'C{row}'].value != None):
        print(Class_1[f'C{row}'])


Class_1_csv = STOP_2001.create_sheet("Class 1 csv") # Create a new sheet that will hold that properly formatted Class 1 data


def yield_excel_cols(cond_col):
    '''
    This function takes a column label and returns the next column.
    It assumes that you need 2 characters at most, e.g. "AF", "BB". This covers 26 + 26^2 = 702 columns.
    We have o.t.o.o. 100 rows to transpose into columns so this should be sufficient.
    To scale this function would require some thought about alphabet arithemtic.
    '''
    new_col = ""
    if(len(cond_col)==1):
        if(str(chr(ord(cond_col[0])+1))=="["):
            new_col = "AA"
        else:
            new_col = str(chr(ord(cond_col[0])+1))
    else:
        if(str(chr(ord(cond_col[-1])+1))=="["):
            new_col = str(chr(ord(cond_col[0])+1)) + "A"
        else:
            new_col = cond_col[0] + str(chr(ord(cond_col[-1])+1))
    
    return new_col

def format_fair_to_new_sheet(col_label, new_row):
    cond_col = "A" # we always start at column A
    for row in range(1,200): # This iterates through the rows of the given fair 
        if (Class_1[f'{col_label}{row}'].value != None and row!=49): # This checks if the row is empty or if it is the secondary name row
            print("row ", row, " : ", Class_1[f'{col_label}{row}'].value) # Print out the values for visual validation
            Class_1_csv[f"{cond_col}{new_row}"] = Class_1[f'{col_label}{row}'].value # Assign the appropriate row value to the appropriate column within the new sheet
            cond_col = yield_excel_cols(cond_col) # Advance the column cursor

def format_feature_labels():
    cond_col="A"
    for row in range(1,200): # We iterate through all the rows to extract the appropriate feature names
        if (Class_1[f'C{row}'].value != None and row!=49):
            Class_1_csv[f"{cond_col}1"] = Class_1[f"A{row}"].value if Class_1[f"A{row}"].value != None else Class_1[f"B{row}"].value
            cond_col = yield_excel_cols(cond_col)
    Class_1_csv["A1"] = "Fair Name"

format_feature_labels()
col = "C"
for row in range(2,30):
    if Class_1[f"{col}2"].value == None:
        break
    format_fair_to_new_sheet(col, row)
    col = yield_excel_cols(col)


STOP_2001.save('2001 STOP Summary.xlsx') # Better save if we want to see our changes

